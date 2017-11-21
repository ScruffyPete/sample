import openpyxl


def export_to_excel(data, headers=None, styles=None):
    if styles is None:
        styles = {}
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    if headers is not None:
        for column_index, header in enumerate(headers, start=1):
            c = worksheet.cell(row=1, column=column_index)
            c.value = header

    for row_index, row in enumerate(data, start=2):
        if isinstance(row, dict):
            row = row.values()
        for column_index, label in enumerate(row, start=1):
            c = worksheet.cell(row=row_index, column=column_index)
            header = headers[column_index-1]
            c.alignment = styles.get(header, {}).get('alignment', None)
            c.value = label

    return workbook
